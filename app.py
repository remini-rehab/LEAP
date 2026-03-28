import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import PatternFill
from fpdf import FPDF
import datetime
import io
import os
import tempfile

# ==========================================
# [0] 화면 기본 설정 및 한글 폰트 로드
# ==========================================
st.set_page_config(page_title="림프부종 ECW 분석 시스템", page_icon="🩺", layout="wide")

# 깃허브에 함께 올린 폰트 파일을 직접 읽어옵니다.
fontpath = "NanumBarunGothic.ttf"
if os.path.exists(fontpath):
    fm.fontManager.addfont(fontpath)
    font_name = fm.FontProperties(fname=fontpath).get_name()
    plt.rc('font', family=font_name)
plt.rc('axes', unicode_minus=False)

# ==========================================
# [1] 보안 로그인 시스템 (비밀번호 체크)
# ==========================================
def check_password():
    """Returns `True` if the user had the correct password."""
    def password_entered():
        # st.secrets에 저장된 비밀번호와 대조
        if st.session_state["password"] == st.secrets["admin_password"]:
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # 보안을 위해 세션에서 지움
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        st.markdown("### 🔒 의료진 전용 시스템입니다.")
        st.text_input("접속 비밀번호를 입력하세요", type="password", on_change=password_entered, key="password")
        return False
    elif not st.session_state["password_correct"]:
        st.markdown("### 🔒 의료진 전용 시스템입니다.")
        st.text_input("비밀번호가 틀렸습니다. 다시 입력하세요", type="password", on_change=password_entered, key="password")
        st.error("비밀번호가 일치하지 않습니다.")
        return False
    return True

# ==========================================
# [2] 데이터 분석 핵심 함수
# ==========================================
@st.cache_data # 웹 속도 향상을 위한 캐싱
def calculate_metrics(df, baseline_days=3):
    df = df.copy()
    df.columns = df.columns.str.strip()
    df['검사일시'] = pd.to_datetime(df['검사일시'], errors='coerce')
    df = df.dropna(subset=['검사일시', '세포외수분비']).sort_values('검사일시')
    df['날짜'] = df['검사일시'].dt.date
    df['시간(Hour)'] = df['검사일시'].dt.hour

    df_am = df[(df['시간(Hour)'] >= 4) & (df['시간(Hour)'] < 14)].groupby('날짜').agg(
        오전_시간=('검사일시', 'first'), 오전_ECW=('세포외수분비', 'first')).reset_index()
    df_pm = df[(df['시간(Hour)'] >= 14) & (df['시간(Hour)'] < 24)].groupby('날짜').agg(
        오후_시간=('검사일시', 'last'), 오후_ECW=('세포외수분비', 'last')).reset_index()

    df_daily = pd.merge(df_am, df_pm, on='날짜', how='outer').sort_values('날짜')
    df_daily = df_daily.rename(columns={'날짜': '검사일시', '오전_ECW': '오전 ECW (A)', '오후_ECW': '오후 ECW (B)'})
    
    if not df_daily.empty:
        full_dates = pd.date_range(start=df_daily['검사일시'].min(), end=df_daily['검사일시'].max()).date
        df_daily = pd.merge(pd.DataFrame({'검사일시': full_dates}), df_daily, on='검사일시', how='left')
    
    df_daily['검사일시'] = pd.to_datetime(df_daily['검사일시'])
    df_daily['day_gain'] = df_daily['오후 ECW (B)'] - df_daily['오전 ECW (A)']
    df_daily['prev_pm'] = df_daily['오후 ECW (B)'].shift(1)
    df_daily['night_recovery'] = df_daily['오전 ECW (A)'] - df_daily['prev_pm']

    baseline_ref = df_daily['오전 ECW (A)'].iloc[:baseline_days].dropna().mean()
    df_daily['baseline_ref'] = baseline_ref
    df_daily['AM_drift'] = df_daily['오전 ECW (A)'] - baseline_ref
    df_daily['AM_3day_range'] = (df_daily['오전 ECW (A)'].rolling(window=3).max() - df_daily['오전 ECW (A)'].rolling(window=3).min()).fillna(0)
    df_daily['recovery_fail'] = np.where(df_daily['night_recovery'].isna(), np.nan, np.where(df_daily['night_recovery'] >= 0, 1, 0))
    df_daily['recovery_fail_3d'] = df_daily['recovery_fail'].rolling(window=3, min_periods=1).sum()
    return df_daily

def generate_detailed_feedback(row):
    if pd.isna(row['오전 ECW (A)']) or pd.isna(row['오후 ECW (B)']):
        return pd.Series(["Level 0 (누락)", "측정 데이터 부족", 0, "-", "-", "-", "-", "-"])

    score = 0
    e_a, e_b, e_c, e_d, e_e = "정상", "정상", "정상", "정상", "정상"

    if not pd.isna(row['AM_drift']):
        if row['AM_drift'] >= 0.004: score += 2; e_a = "상승 심화 (+2점)"
        elif row['AM_drift'] >= 0.002: score += 1; e_a = "미세 상승 (+1점)"
    if not pd.isna(row['day_gain']):
        if row['day_gain'] >= 0.008: score += 2; e_b = "축적 심화 (+2점)"
        elif row['day_gain'] >= 0.005: score += 1; e_b = "축적 증가 (+1점)"
    if not pd.isna(row['night_recovery']):
        if row['night_recovery'] >= 0: score += 2; e_c = "회복 실패 (+2점)"
        elif row['night_recovery'] > -0.002: score += 1; e_c = "회복 미흡 (+1점)"
    if not pd.isna(row['AM_3day_range']):
        if row['AM_3day_range'] >= 0.006: score += 2; e_d = "변동폭 심화 (+2점)"
        elif row['AM_3day_range'] >= 0.003: score += 1; e_d = "변동성 증가 (+1점)"
    if not pd.isna(row['recovery_fail_3d']):
        if row['recovery_fail_3d'] >= 2: score += 2; e_e = "회복실패 반복 (+2점)"

    if score >= 4 or (not pd.isna(row['recovery_fail_3d']) and row['recovery_fail_3d'] == 3):
        level = "Level 3 (경고)"; msg = "의료진 개입 필요"
    elif score >= 2:
        level = "Level 2 (주의)"; msg = "자가관리 점검"
    else:
        level = "Level 1 (안정)"; msg = "현재 유지"

    return pd.Series([level, msg, score, e_a, e_b, e_c, e_d, e_e])

# PDF 클래스
class PatientPDF(FPDF):
    def header(self):
        self.add_font("Nanum", "", fontpath)
        self.set_font("Nanum", "", 16)
        self.cell(0, 15, "환자별 주간 체수분 분석 보고서", ln=True, align="C")
        self.ln(5)

# ==========================================
# [3] 메인 웹 대시보드 화면
# ==========================================
if check_password(): # 로그인이 성공해야 아래 화면이 보입니다.
    st.title("🩺 스마트 ECW 부종 분석 센터")
    st.markdown("인바디 엑셀 데이터를 업로드하면 즉시 분석 결과와 리포트를 제공합니다.")

    uploaded_file = st.sidebar.file_uploader("📂 환자 엑셀 파일 업로드", type=["xlsx"])
    
    if uploaded_file is not None:
        with st.spinner("데이터를 분석하고 리포트를 생성하는 중입니다..."):
            multi_sheet_data = pd.read_excel(uploaded_file, sheet_name=None)
            eval_cols = ['판정 수준', '메시지', '총점', '기저선', '주간축적', '야간회복', '변동성', '회복반복']
            
            # 통합 엑셀 다운로드를 위한 메모리 버퍼
            excel_buffer = io.BytesIO()
            
            # 탭(Tab)으로 환자별 화면을 예쁘게 분리
            patient_names = list(multi_sheet_data.keys())
            tabs = st.tabs(patient_names)
            
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                for tab, patient_name in zip(tabs, patient_names):
                    patient_df = multi_sheet_data[patient_name]
                    
                    with tab:
                        if patient_df.empty:
                            st.warning(f"{patient_name} 환자의 데이터가 비어 있습니다.")
                            continue
                            
                        # 분석 실행
                        df_analyzed = calculate_metrics(patient_df)
                        df_analyzed[eval_cols] = df_analyzed.apply(generate_detailed_feedback, axis=1)
                        df_analyzed['오전_시간'] = pd.to_datetime(df_analyzed['오전_시간'])
                        df_analyzed['오후_시간'] = pd.to_datetime(df_analyzed['오후_시간'])
                        
                        this_week = df_analyzed.tail(7)
                        tw_warnings = len(this_week[this_week['판정 수준'].str.contains('Level 2|Level 3')])
                        
                        st.subheader(f"👤 {patient_name} 환자 요약")
                        col1, col2, col3 = st.columns(3)
                        col1.metric("이번 주 경고/주의 횟수", f"{tw_warnings} 건")
                        col2.metric("최근 상태", this_week['판정 수준'].iloc[-1] if not this_week.empty else "데이터 없음")
                        col3.metric("야간 회복률", "확인 요망" if tw_warnings > 0 else "안정적")
                        
                        # 그래프 생성 (화면 출력용)
                        fig, ax = plt.subplots(figsize=(10, 4))
                        x_dates = df_analyzed['검사일시'].dt.strftime('%m-%d')
                        ax.plot(x_dates, df_analyzed['오전 ECW (A)'], 'b-o', label='오전 기저선')
                        ax.plot(x_dates, df_analyzed['오후 ECW (B)'], 'r--x', label='오후 축적량')
                        ax.set_title(f"{patient_name} ECW Trend")
                        ax.legend()
                        ax.grid(alpha=0.3)
                        st.pyplot(fig) # 웹 화면에 그래프 그리기
                        
                        # PDF 생성을 위한 임시 이미지 저장 (서버 찌꺼기 방지용 tempfile)
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmpfile:
                            fig.savefig(tmpfile.name)
                            img_path = tmpfile.name
                            
                        # 엑셀 시트 저장
                        df_analyzed.to_excel(writer, sheet_name=patient_name, index=False)
                        worksheet = writer.sheets[patient_name]
                        worksheet.add_image(OpenpyxlImage(img_path), 'O1')
                        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                        for cell in worksheet["1:1"]: cell.fill = header_fill
                            
                        # PDF 생성 (메모리상에서)
                        pdf = PatientPDF()
                        pdf.add_font("Nanum", "", fontpath)
                        pdf.add_page()
                        pdf.set_font("Nanum", "", 12)
                        pdf.cell(0, 10, f"환자명: {patient_name}  |  분석일: {datetime.date.today()}", ln=True)
                        pdf.image(img_path, x=10, w=190)
                        pdf.ln(5)
                        pdf.set_font("Nanum", "", 9)
                        pdf.cell(0, 10, "[최근 5일 상세 데이터]", ln=True)
                        for _, r in this_week.tail(5).iterrows():
                            pdf.cell(0, 7, f"{r['검사일시'].strftime('%Y-%m-%d')}: {r['판정 수준']} | 오전:{r['오전 ECW (A)']:.3f} | 오후:{r['오후 ECW (B)']:.3f}", ln=True)
                        
                        # 개별 PDF 다운로드 버튼
                        pdf_bytes = bytes(pdf.output())
                        st.download_button(
                            label=f"📥 {patient_name} PDF 리포트 다운로드",
                            data=pdf_bytes,
                            file_name=f"리포트_{patient_name}.pdf",
                            mime="application/pdf",
                            key=f"pdf_{patient_name}"
                        )
                        
                        st.markdown("---")
            
            # 전체 분석 엑셀 다운로드 (사이드바)
            st.sidebar.success("분석 완료!")
            st.sidebar.download_button(
                label="📊 통합 엑셀 다운로드 (전체 환자)",
                data=excel_buffer.getvalue(),
                file_name=f"분석결과_통합_{datetime.date.today().strftime('%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.info("👈 왼쪽 사이드바에서 분석할 엑셀 파일을 업로드해 주세요.")